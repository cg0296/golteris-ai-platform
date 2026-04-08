"""
backend/services/email_send.py — Outbound email send service (#25).

This is the C2 enforcement point for outbound communication. It verifies
that an approval has status=APPROVED before sending any email. Nothing
bypasses this check — not even in tests.

Flow:
    1. Worker dispatches a "send_outbound_email" job with {approval_id}
    2. This service loads the approval, verifies status == APPROVED
    3. Gets the email body (resolved_body if edited, otherwise draft_body)
    4. Sends via the configured email provider (Graph, IMAP, or file)
    5. Persists an outbound Message row for the RFQ thread
    6. Creates an AuditEvent ("email_sent")
    7. On failure: logs error, creates a review AuditEvent (FR-HI-6)

Cross-cutting constraints:
    C2 — No email sends without approved=true on the draft record
    C4 — Every send/failure is audited with full traceability
    FR-HI-1 — Drafts persist with pending_approval; only send after approved
    FR-HI-6 — Failed sends create a review card, not silent disappearance

Called by:
    backend/worker.py via JOB_DISPATCH["send_outbound_email"]
"""

import logging
import os
from datetime import datetime, timezone

from sqlalchemy.orm import Session

from backend.db.models import (
    Approval,
    ApprovalStatus,
    AuditEvent,
    Message,
    MessageDirection,
)
from backend.services.email_ingestion import get_provider_from_config

logger = logging.getLogger("golteris.services.email_send")


def send_approved_email(db: Session, approval_id: int) -> None:
    """
    Send an outbound email for an approved draft.

    This is the ONLY function that sends email. It is called by the worker
    when processing a "send_outbound_email" job.

    C2 HARD GATE: If approval.status != APPROVED, this function refuses
    to send and logs a warning. This is the final safety check — even if
    a bug enqueues a send job for a non-approved draft, it will not send.

    Args:
        db: SQLAlchemy session
        approval_id: ID of the approved Approval record

    Side effects:
        - Sends an email via the configured provider
        - Creates an outbound Message row linked to the RFQ
        - Creates an AuditEvent for traceability (C4)
        - On failure: creates a failure AuditEvent (FR-HI-6)
    """
    # Load the approval
    approval = db.query(Approval).filter(Approval.id == approval_id).first()
    if not approval:
        logger.error("Send job for approval %d: not found", approval_id)
        return

    # ===== C2 ENFORCEMENT — THE CRITICAL CHECK =====
    # If the approval is not in APPROVED status, refuse to send.
    # This is the absolute last line of defense. Even if a send job
    # was enqueued by mistake, this check prevents unauthorized sends.
    if approval.status != ApprovalStatus.APPROVED:
        logger.warning(
            "Send job for approval %d: status is %s, not APPROVED — refusing to send (C2)",
            approval_id,
            approval.status.value,
        )
        return

    # Determine which body to send — edited version if broker edited, otherwise original
    email_body = approval.resolved_body or approval.draft_body
    email_subject = approval.draft_subject or "(no subject)"
    email_to = approval.draft_recipient

    # Inject RFQ reference tag into the subject so replies carry it back.
    # This gives the matching service a deterministic way to link replies
    # to the correct RFQ, even when thread headers are lost or the sender
    # has multiple active RFQs. Format: [RFQ-42] at the end of the subject.
    if approval.rfq_id and f"[RFQ-{approval.rfq_id}]" not in email_subject:
        email_subject = f"{email_subject} [RFQ-{approval.rfq_id}]"

    if not email_to:
        logger.error("Send job for approval %d: no recipient", approval_id)
        _create_failure_event(
            db, approval, "No recipient address on the draft"
        )
        return

    # Find the original inbound message for threading (In-Reply-To header)
    reply_to_message_id = None
    if approval.rfq_id:
        original = (
            db.query(Message)
            .filter(
                Message.rfq_id == approval.rfq_id,
                Message.direction == MessageDirection.INBOUND,
            )
            .order_by(Message.received_at.desc())
            .first()
        )
        if original:
            reply_to_message_id = original.message_id_header

    # Check if a quote sheet should be attached (#152)
    attachment = None
    if approval.reason and "[ATTACH_QUOTE_SHEET]" in approval.reason and approval.rfq_id:
        try:
            attachment = _generate_quote_sheet_attachment(db, approval.rfq_id)
            logger.info("Quote sheet attachment generated for approval %d", approval_id)
        except Exception as e:
            logger.warning("Could not generate quote sheet attachment: %s", e)

    # Send via the configured provider
    provider = get_provider_from_config()
    logger.info(
        "Sending email via %s: to=%s subject=%s (approval=%d)",
        provider.get_provider_name(), email_to, email_subject, approval_id,
    )

    result = provider.send_message(
        to=email_to,
        subject=email_subject,
        body=email_body,
        reply_to_message_id=reply_to_message_id,
        attachment=attachment,
    )

    if result["success"]:
        _handle_send_success(db, approval, email_to, email_subject, email_body, result)
    else:
        _handle_send_failure(db, approval, result["error"])


def _handle_send_success(
    db: Session,
    approval: Approval,
    to: str,
    subject: str,
    body: str,
    result: dict,
) -> None:
    """
    After a successful send: persist the outbound message and audit event.

    Creates a Message row with direction=OUTBOUND so it appears in the
    RFQ's message thread. Creates an AuditEvent so the broker sees
    "Email sent to [recipient]" in the timeline.
    """
    # Persist the outbound message in the RFQ thread.
    # Use the broker's name + sending address for the sender field.
    broker_name = _get_broker_name(db)
    send_address = os.environ.get("MS_GRAPH_USER_EMAIL", "agents@golteris.com")
    outbound_msg = Message(
        rfq_id=approval.rfq_id,
        direction=MessageDirection.OUTBOUND,
        sender=f"{broker_name} <{send_address}>",
        recipients=to,
        subject=subject,
        body=body,
        received_at=datetime.now(timezone.utc),
    )
    db.add(outbound_msg)

    # Audit event — the broker sees "Email sent" in the timeline (C4)
    type_label = approval.approval_type.value.replace("_", " ")
    event = AuditEvent(
        rfq_id=approval.rfq_id,
        event_type="email_sent",
        actor="system",
        description=f"Sent {type_label} to {to}",
        event_data={
            "approval_id": approval.id,
            "approval_type": approval.approval_type.value,
            "recipient": to,
            "subject": subject,
            "provider_message_id": result.get("message_id"),
        },
    )
    db.add(event)
    db.commit()

    logger.info("Email sent successfully for approval %d", approval.id)


def _generate_quote_sheet_attachment(db: Session, rfq_id: int) -> dict | None:
    """
    Generate an Excel quote sheet as an attachment dict (#152).

    Returns a dict with filename, content_type, and base64-encoded data
    that the email provider can attach to the outbound email.
    """
    import base64
    from io import BytesIO

    # Reuse the download endpoint logic to generate the Excel bytes
    from backend.api.carriers import get_quote_sheet
    from backend.db.models import RFQ

    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    if not rfq:
        return None

    try:
        sheet_response = get_quote_sheet(rfq_id, db)
    except Exception:
        return None

    sheet = sheet_response["quote_sheet"]
    lanes = sheet.get("lanes", [])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote Sheet"

    # Minimal styling for the attachment
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    col_header_fill = PatternFill(start_color="0E2841", end_color="0E2841", fill_type="solid")
    col_header_font = Font(bold=True, size=11, color="FFFFFF")

    ref_id = sheet.get("reference_id", f"RFQ-{rfq_id}")
    ws["A1"] = "Carrier Quote Request"
    ws["A1"].font = Font(bold=True, size=14)

    from datetime import datetime
    details = [
        ("Reference:", ref_id),
        ("Equipment:", rfq.equipment_type or "—"),
        ("Commodity:", rfq.commodity or "—"),
        ("Special:", sheet.get("special_requirements", rfq.special_requirements or "None")),
    ]
    for row_idx, (label, value) in enumerate(details, 3):
        ws.cell(row=row_idx, column=1, value=label).font = header_font
        ws.cell(row=row_idx, column=2, value=str(value))

    # Lane table
    table_row = len(details) + 4
    col_headers = ["Lane", "Origin", "Destination", "Commodity", "Weight (lbs)", "# Trucks", "Rate / Amount", "Available (Y/N)"]
    for col_idx, h in enumerate(col_headers, 1):
        cell = ws.cell(row=table_row, column=col_idx, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.border = thin_border

    for lane_idx, lane in enumerate(lanes, 1):
        r = table_row + lane_idx
        vals = [lane_idx, lane.get("origin", ""), lane.get("destination", ""),
                lane.get("commodity", ""), lane.get("weight_lbs", ""),
                lane.get("truck_count", ""), "", ""]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.border = thin_border

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return {
        "filename": f"{ref_id.replace(' ', '_')}_quote_sheet.xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "data_base64": base64.b64encode(buf.read()).decode("utf-8"),
    }


def _get_broker_name(db: Session) -> str:
    """Get the active broker's first name for outbound message records."""
    try:
        from backend.db.models import User
        user = db.query(User).filter(User.active == True).order_by(User.id.desc()).first()
        if user and user.name:
            return user.name.split()[0]
    except Exception:
        pass
    return "Beltmann Logistics"


def _handle_send_failure(
    db: Session,
    approval: Approval,
    error: str,
) -> None:
    """
    After a failed send: create a review card so the failure is visible (FR-HI-6).

    Failed sends do NOT silently disappear. The broker sees a failure event
    in the timeline and can investigate or retry.
    """
    _create_failure_event(db, approval, error)
    logger.error("Email send failed for approval %d: %s", approval.id, error)


def _create_failure_event(
    db: Session,
    approval: Approval,
    error: str,
) -> None:
    """
    Create an audit event for a send failure (FR-HI-6).

    The broker sees "Failed to send [type] to [recipient]" in the timeline
    with the error details in event_data.
    """
    type_label = approval.approval_type.value.replace("_", " ")
    recipient = approval.draft_recipient or "unknown"
    event = AuditEvent(
        rfq_id=approval.rfq_id,
        event_type="email_send_failed",
        actor="system",
        description=f"Failed to send {type_label} to {recipient}: {error}",
        event_data={
            "approval_id": approval.id,
            "approval_type": approval.approval_type.value,
            "recipient": recipient,
            "error": error,
        },
    )
    db.add(event)
    db.commit()
