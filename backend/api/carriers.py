"""
backend/api/carriers.py — FastAPI router for carrier management (#32).

Endpoints:
    GET  /api/carriers              — List all active carriers
    GET  /api/carriers/match/{rfq_id} — Carriers matching an RFQ's equipment/lane
    POST /api/rfqs/{rfq_id}/distribute — Distribute RFQ to selected carriers

Cross-cutting constraints:
    C2 — Distribution creates a pending approval; sends only after approval
"""

import logging
from typing import List

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session

from backend.db.database import get_db
from backend.db.models import Carrier, RFQ
from backend.services.carrier_distribution import (
    distribute_to_carriers,
    get_matching_carriers,
    list_carriers,
)

logger = logging.getLogger("golteris.api.carriers")

router = APIRouter(tags=["carriers"])


@router.get("/api/carriers")
def get_carriers(db: Session = Depends(get_db)):
    """List all active carriers for the carrier selection UI."""
    carriers = list_carriers(db, active_only=True)
    return {
        "carriers": [_serialize_carrier(c) for c in carriers],
        "total": len(carriers),
    }


@router.post("/api/carriers")
def create_carrier(body: dict, db: Session = Depends(get_db)):
    """Create a new carrier (#137)."""
    carrier = Carrier(
        name=body.get("name", ""),
        email=body.get("email", ""),
        contact_name=body.get("contact_name", ""),
        phone=body.get("phone", ""),
        equipment_types=body.get("equipment_types", []),
        lanes=body.get("lanes", []),
        preferred=body.get("preferred", False),
    )
    db.add(carrier)
    db.commit()
    db.refresh(carrier)
    return _serialize_carrier(carrier)


@router.patch("/api/carriers/{carrier_id}")
def update_carrier(carrier_id: int, body: dict, db: Session = Depends(get_db)):
    """Update a carrier (#137)."""
    carrier = db.query(Carrier).filter(Carrier.id == carrier_id).first()
    if not carrier:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Carrier not found")
    for key in ["name", "email", "contact_name", "phone", "equipment_types", "lanes", "preferred"]:
        if key in body:
            setattr(carrier, key, body[key])
    db.commit()
    db.refresh(carrier)
    return _serialize_carrier(carrier)


@router.delete("/api/carriers/{carrier_id}")
def delete_carrier(carrier_id: int, db: Session = Depends(get_db)):
    """Delete a carrier (#137)."""
    carrier = db.query(Carrier).filter(Carrier.id == carrier_id).first()
    if not carrier:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Carrier not found")
    db.delete(carrier)
    db.commit()
    return {"status": "ok"}


@router.get("/api/carriers/match/{rfq_id}")
def get_matching(rfq_id: int, db: Session = Depends(get_db)):
    """
    Return carriers matching an RFQ's equipment type and lane.

    Used by the carrier selection modal to pre-check matching carriers.
    """
    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    if not rfq:
        raise HTTPException(status_code=404, detail=f"RFQ {rfq_id} not found")

    carriers = get_matching_carriers(db, rfq)
    return {
        "carriers": [_serialize_carrier(c) for c in carriers],
        "total": len(carriers),
        "rfq_id": rfq_id,
    }


class DistributeRequest(BaseModel):
    """Request body for carrier RFQ distribution."""
    carrier_ids: List[int]
    attach_quote_sheet: bool = False


class PriceRequest(BaseModel):
    """Request body for pricing an RFQ with a selected carrier bid."""
    carrier_bid_id: int
    manual_rate: float | None = None
    override_reason: str | None = None


@router.post("/api/rfqs/{rfq_id}/redraft")
def redraft_rfq(rfq_id: int, db: Session = Depends(get_db)):
    """
    Trigger the AI to generate a new draft for this RFQ.

    Used when the broker rejected the previous draft and wants a fresh one.
    Calls the validation agent to draft a new follow-up.
    """
    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    if not rfq:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="RFQ not found")

    try:
        from backend.agents.validation import draft_followup
        result = draft_followup(db, rfq_id)
        return {"status": "ok", "message": "New draft generated — check Urgent Actions", "result": result}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@router.post("/api/rfqs/{rfq_id}/manual-reply")
def manual_reply(rfq_id: int, body: dict, db: Session = Depends(get_db)):
    """
    Send a manual reply for this RFQ (broker-written, not AI-drafted).

    Creates an approval record with the broker's text, auto-approves it,
    and queues for sending. C2: still goes through the send pipeline.
    """
    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    if not rfq:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="RFQ not found")

    to = body.get("to", rfq.customer_email or "")
    subject = body.get("subject", f"Re: {rfq.customer_name or 'Quote Request'}")
    email_body = body.get("body", "")

    if not to or not email_body:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="'to' and 'body' are required")

    # Create an auto-approved approval for the manual reply
    from backend.db.models import Approval, ApprovalType, ApprovalStatus
    approval = Approval(
        rfq_id=rfq_id,
        approval_type=ApprovalType.CUSTOMER_REPLY,
        draft_body=email_body,
        draft_subject=subject,
        draft_recipient=to,
        reason="Manual reply by broker",
        status=ApprovalStatus.APPROVED,
        approved_by="operator",
    )
    db.add(approval)
    db.commit()

    # Enqueue the send job
    from backend.worker import enqueue_job
    enqueue_job(db, "send_outbound_email", {"approval_id": approval.id})

    return {"status": "ok", "message": f"Reply queued for sending to {to}", "approval_id": approval.id}


@router.post("/api/rfqs/{rfq_id}/distribute")
def distribute_rfq(
    rfq_id: int,
    body: DistributeRequest,
    db: Session = Depends(get_db),
):
    """
    Distribute an RFQ to selected carriers.

    Creates a batch approval (C2 gate) and per-carrier send tracking.
    The broker must approve before any emails are sent.
    """
    try:
        result = distribute_to_carriers(db, rfq_id, body.carrier_ids, attach_quote_sheet=body.attach_quote_sheet)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return result


@router.post("/api/rfqs/{rfq_id}/price")
def price_rfq(
    rfq_id: int,
    body: PriceRequest,
    db: Session = Depends(get_db),
):
    """
    Apply markup to a selected carrier bid and set the customer-facing rate (#35).

    If manual_rate is provided, uses the broker's override (audited).
    Otherwise applies the default 12% markup with $150 minimum margin.
    """
    from decimal import Decimal
    from backend.services.pricing import calculate_customer_rate

    try:
        result = calculate_customer_rate(
            db, rfq_id,
            carrier_bid_id=body.carrier_bid_id,
            manual_rate=Decimal(str(body.manual_rate)) if body.manual_rate else None,
            override_reason=body.override_reason,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "carrier_bid_id": result.carrier_bid_id,
        "carrier_name": result.carrier_name,
        "carrier_rate": float(result.carrier_rate),
        "markup_percent": float(result.markup_percent),
        "markup_amount": float(result.markup_amount),
        "customer_rate": float(result.customer_rate),
        "margin": float(result.margin),
        "is_manual_override": result.is_manual_override,
    }


@router.post("/api/rfqs/{rfq_id}/outcome")
def set_rfq_outcome(
    rfq_id: int,
    body: dict,
    db: Session = Depends(get_db),
):
    """
    Mark an RFQ as Won, Lost, or Cancelled (#100).

    Transitions the RFQ to the terminal state, sets the outcome field,
    records closed_at, and creates an audit event.
    """
    from backend.services.rfq_state_machine import transition_rfq
    from backend.db.models import RFQState, AuditEvent
    from datetime import datetime, timezone

    outcome = body.get("outcome", "").lower()
    if outcome not in ("won", "lost", "cancelled"):
        raise HTTPException(status_code=400, detail=f"Invalid outcome: {outcome}. Must be won, lost, or cancelled.")

    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    if not rfq:
        raise HTTPException(status_code=404, detail=f"RFQ {rfq_id} not found")

    state_map = {"won": RFQState.WON, "lost": RFQState.LOST, "cancelled": RFQState.CANCELLED}
    target_state = state_map[outcome]

    try:
        transition_rfq(db, rfq_id, target_state, actor="broker", reason=body.get("reason", f"Marked as {outcome}"))
    except Exception:
        # Use override if normal transition isn't allowed from current state
        from backend.services.rfq_state_machine import override_rfq_state
        override_rfq_state(db, rfq_id, target_state, actor="broker", reason=body.get("reason", f"Marked as {outcome}"))

    # Set outcome fields
    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    rfq.outcome = outcome
    rfq.closed_at = datetime.now(timezone.utc)
    db.commit()

    return {"id": rfq_id, "state": outcome, "closed_at": rfq.closed_at.isoformat()}


@router.get("/api/rfqs/{rfq_id}/quote-sheet")
def get_quote_sheet(rfq_id: int, db: Session = Depends(get_db)):
    """
    Return the generated quote sheet for an RFQ.

    Reads the quote_sheet agent's tool-use output from agent_calls
    and formats it for display or download.
    """
    from backend.db.models import AgentCall, AgentRun
    import json

    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    if not rfq:
        raise HTTPException(status_code=404, detail=f"RFQ {rfq_id} not found")

    # Find the most recent quote_sheet agent call for this RFQ
    call = (
        db.query(AgentCall)
        .join(AgentRun, AgentCall.run_id == AgentRun.id)
        .filter(AgentRun.rfq_id == rfq_id, AgentCall.agent_name == "quote_sheet")
        .order_by(AgentCall.started_at.desc())
        .first()
    )

    if not call or not call.response:
        raise HTTPException(status_code=404, detail="No quote sheet found for this RFQ")

    # Parse the stored response. New format: clean JSON dict from the agent.
    # Legacy format: str(raw_response) from the Anthropic SDK — needs extraction.
    try:
        sheet_data = json.loads(call.response)
        # Verify it looks like a quote sheet (has expected keys)
        if not isinstance(sheet_data, dict) or "lanes" not in sheet_data:
            raise ValueError("Not a quote sheet dict")
    except (json.JSONDecodeError, ValueError):
        # Legacy fallback: parse str(raw_response) with input= extraction
        try:
            import ast
            resp = call.response
            idx = resp.find("input=")
            if idx >= 0:
                start = resp.index("{", idx)
                depth = 0
                end = start
                for i in range(start, len(resp)):
                    if resp[i] == "{":
                        depth += 1
                    elif resp[i] == "}":
                        depth -= 1
                        if depth == 0:
                            end = i + 1
                            break
                input_str = resp[start:end]
                sheet_data = ast.literal_eval(input_str)
            else:
                sheet_data = {"raw": resp[:2000]}
        except Exception:
            sheet_data = {"raw": call.response[:2000]}

    return {
        "rfq_id": rfq_id,
        "customer_name": rfq.customer_name,
        "customer_company": rfq.customer_company,
        "quote_sheet": sheet_data,
    }


@router.post("/api/rfqs/{rfq_id}/regenerate-quote-sheet")
def regenerate_quote_sheet(rfq_id: int, db: Session = Depends(get_db)):
    """
    Re-run the quote sheet agent for an RFQ (#156).

    Temporarily sets the RFQ to ready_to_quote so the agent will process it,
    then restores the original state. This allows regeneration from any state.
    """
    from backend.agents.quote_sheet import generate_quote_sheet
    from backend.db.models import AuditEvent, RFQState

    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    if not rfq:
        raise HTTPException(status_code=404, detail=f"RFQ {rfq_id} not found")

    original_state = rfq.state
    # Temporarily set to ready_to_quote so the agent will run
    if rfq.state != RFQState.READY_TO_QUOTE:
        rfq.state = RFQState.READY_TO_QUOTE
        db.commit()

    try:
        sheet = generate_quote_sheet(db, rfq_id)
    finally:
        # Restore original state
        if rfq.state != original_state:
            rfq.state = original_state
            db.commit()

    if not sheet:
        raise HTTPException(status_code=500, detail="Quote sheet generation failed")

    db.add(AuditEvent(
        rfq_id=rfq_id,
        event_type="quote_sheet_regenerated",
        actor="broker",
        description="Quote sheet regenerated manually",
    ))
    db.commit()

    return {"status": "ok", "rfq_id": rfq_id, "quote_sheet": sheet}


@router.get("/api/rfqs/{rfq_id}/quote-sheet/download")
def download_quote_sheet_excel(rfq_id: int, db: Session = Depends(get_db)):
    """
    Download the quote sheet as an Excel (.xlsx) file for carrier distribution (#144).

    Header section: reference, dates, lane count, equipment, commodity, special instructions.
    Lane table: origin, dest, commodity, weight, trucks, plus empty Rate and Available columns.
    """
    from io import BytesIO
    from datetime import datetime
    from fastapi.responses import StreamingResponse

    # Reuse the existing quote-sheet endpoint logic to get the data
    sheet_response = get_quote_sheet(rfq_id, db)
    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    sheet = sheet_response["quote_sheet"]
    lanes = sheet.get("lanes", [])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote Sheet"

    # -- Styles --
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    col_header_font = Font(bold=True, size=11, color="FFFFFF")
    col_header_fill = PatternFill(start_color="0E2841", end_color="0E2841", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # -- Header Section --
    ws.merge_cells("A1:H1")
    ws["A1"] = "Carrier Quote Request"
    ws["A1"].font = Font(bold=True, size=16)

    ref_id = sheet.get("reference_id", f"RFQ-{rfq_id}")
    details = [
        ("Reference:", ref_id),
        ("Generated:", datetime.utcnow().strftime("%B %d, %Y")),
        ("Requested Due:", sheet.get("response_deadline", "ASAP")),
        ("Number of Lanes:", len(lanes)),
        ("Equipment:", rfq.equipment_type or "—"),
        ("Commodity:", rfq.commodity or "—"),
        ("Special Instructions:", sheet.get("special_requirements", rfq.special_requirements or "None")),
    ]

    row = 3
    for label, value in details:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=str(value)).font = value_font
        row += 1

    # Summary line
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1, value=sheet.get("summary", "")).font = Font(italic=True, size=11)

    # -- Lane Table --
    row += 2
    col_headers = ["Lane", "Origin", "Destination", "Commodity", "Weight (lbs)", "# Trucks", "Rate / Amount", "Available (Y/N)"]
    col_widths = [8, 22, 22, 18, 15, 12, 16, 16]

    for col_idx, header in enumerate(col_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = col_widths[col_idx - 1]

    row += 1
    for lane_idx, lane in enumerate(lanes, 1):
        values = [
            lane_idx,
            lane.get("origin", ""),
            lane.get("destination", ""),
            lane.get("commodity", ""),
            lane.get("weight_lbs", ""),
            lane.get("truck_count", ""),
            "",  # Rate — empty for carrier to fill
            "",  # Available — empty for carrier to fill
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = value_font
            if col_idx in (1, 5, 6):
                cell.alignment = Alignment(horizontal="center")
            # Highlight the empty input columns with a light yellow background
            if col_idx >= 7:
                cell.fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        row += 1

    # Notes row
    if sheet.get("notes"):
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value=f"Notes: {sheet['notes']}").font = Font(italic=True, size=10)

    # Write to buffer and return
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{ref_id.replace(' ', '_')}_quote_sheet.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/api/rfqs/{rfq_id}/generate-quote")
def generate_quote(
    rfq_id: int,
    db: Session = Depends(get_db),
):
    """
    Generate a customer-facing quote and create a pending approval (#36).

    Uses the RFQ's quoted_amount (set by pricing engine #35) to generate
    a professional email template. The broker reviews in the approval modal
    before sending (C2 gate).
    """
    from backend.services.customer_quote import generate_customer_quote

    try:
        result = generate_customer_quote(db, rfq_id)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return result


@router.get("/api/rfqs/{rfq_id}/bids")
def get_ranked_bids(rfq_id: int, db: Session = Depends(get_db)):
    """
    Return ranked carrier bids for an RFQ (#34).

    Bids are sorted by total landed cost with tags:
    best_value, runner_up, outlier_high, outlier_low.
    """
    from backend.services.bid_ranking import rank_bids

    rfq = db.query(RFQ).filter(RFQ.id == rfq_id).first()
    if not rfq:
        raise HTTPException(status_code=404, detail=f"RFQ {rfq_id} not found")

    ranked = rank_bids(db, rfq_id)
    return {
        "rfq_id": rfq_id,
        "bids": [
            {
                "id": rb.bid.id,
                "rank": rb.rank,
                "carrier_name": rb.bid.carrier_name,
                "carrier_email": rb.bid.carrier_email,
                "rate": float(rb.bid.rate) if rb.bid.rate else None,
                "normalized_rate": rb.normalized_rate,
                "currency": rb.bid.currency,
                "rate_type": rb.bid.rate_type,
                "terms": rb.bid.terms,
                "availability": rb.bid.availability,
                "notes": rb.bid.notes,
                "tag": rb.tag,
                "reason": rb.reason,
                "received_at": rb.bid.received_at.isoformat() if rb.bid.received_at else None,
            }
            for rb in ranked
        ],
        "total": len(ranked),
    }


def _serialize_carrier(carrier: Carrier) -> dict:
    """Convert a Carrier to a JSON dict for the selection UI."""
    return {
        "id": carrier.id,
        "name": carrier.name,
        "email": carrier.email,
        "contact_name": carrier.contact_name,
        "phone": carrier.phone,
        "equipment_types": carrier.equipment_types or [],
        "lanes": carrier.lanes or [],
        "preferred": carrier.preferred,
    }
